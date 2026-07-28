from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from datetime import datetime, date
import os
import openpyxl
import json
import re

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'iim-sambalpur-timetable-secret-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timetable.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
migrate = Migrate(app, db)

# ─── MODELS ────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)


class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    credits = db.Column(db.Float, default=3.0)
    area = db.Column(db.String(50))
    faculty = db.Column(db.String(300))
    short_name = db.Column(db.String(20))  # e.g., "FRA", "MM", "DS-I"
    color = db.Column(db.String(7), default='#6366f1')  # hex color

    def to_dict(self):
        return {
            'id': self.id,
            'code': self.code,
            'name': self.name,
            'credits': self.credits,
            'area': self.area,
            'faculty': self.faculty,
            'short_name': self.short_name,
            'color': self.color
        }


class ClassSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    day_name = db.Column(db.String(20))
    slot = db.Column(db.Integer, nullable=False)  # 1=9:30AM, 2=11:30AM, 3=2:00PM, 4=4:00PM
    subject_raw = db.Column(db.String(300))       # raw text from Excel / edited
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=True)
    course = db.relationship('Course', backref='sessions')
    is_special = db.Column(db.Boolean, default=False)  # holiday / exam
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        slot_times = {1: '09:30 AM', 2: '11:30 AM', 3: '02:00 PM', 4: '04:00 PM'}
        return {
            'id': self.id,
            'date': self.date.isoformat(),
            'day_name': self.day_name,
            'slot': self.slot,
            'slot_time': slot_times.get(self.slot, ''),
            'subject_raw': self.subject_raw,
            'course': self.course.to_dict() if self.course else None,
            'is_special': self.is_special,
            'notes': self.notes
        }


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='notifications')
    push_subscription = db.Column(db.Text)  # JSON Web Push subscription
    notify_before_class = db.Column(db.Boolean, default=True)
    notify_minutes_before = db.Column(db.Integer, default=15)
    notify_morning = db.Column(db.Boolean, default=True)
    morning_time = db.Column(db.String(5), default='07:00')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ─── SLOT TIME CONSTANTS ───────────────────────────────────────────────────────
SLOTS = {
    1: {'label': '09:30 AM – 11:00 AM', 'start': '09:30', 'end': '11:00'},
    2: {'label': '11:30 AM – 01:00 PM', 'start': '11:30', 'end': '13:00'},
    3: {'label': '02:00 PM – 03:30 PM', 'start': '14:00', 'end': '15:30'},
    4: {'label': '04:00 PM – 05:30 PM', 'start': '16:00', 'end': '17:30'},
}

COURSE_COLORS = [
    '#6366f1', '#8b5cf6', '#ec4899', '#f59e0b',
    '#10b981', '#3b82f6', '#ef4444', '#14b8a6',
]

COURSE_ABBR_MAP = {
    'FRA': 'FRA', 'MM': 'MM', 'DS': 'DS-I', 'DS-I': 'DS-I',
    'MC': 'MC', 'MComp': 'MComp', 'OBD': 'OBD', 'BE': 'BE',
    'IBA': 'IBA',
}


def parse_subject_abbr(text):
    """Extract short course code from cell text."""
    if not text:
        return None
    text = text.strip()
    # Try to find known abbreviations
    for abbr in ['DS-I', 'MComp', 'FRA', 'MM', 'MC', 'OBD', 'BE', 'IBA']:
        if text.upper().startswith(abbr.upper()) or f'{abbr}-' in text or f'{abbr} ' in text.upper():
            return abbr
    # Fallback: first word
    return text.split()[0] if text else None


def import_excel(filepath):
    """Parse Excel timetable and populate DB."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Build course map from footer rows
    course_color_idx = 0
    courses_by_short = {}

    # Read course table at bottom (after row 60 roughly)
    for row in ws.iter_rows(min_row=60, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        try:
            no = float(row[0])
        except (TypeError, ValueError):
            continue
        name = str(row[1]).strip() if row[1] else None
        credits = float(row[2]) if row[2] else 3.0
        area = str(row[3]).strip() if row[3] else None
        code = str(row[4]).strip() if row[4] else f'MBA-BA{int(no):03d}'
        faculty = str(row[5]).strip() if row[5] else None

        if not name:
            continue

        # Determine short name
        short_map = {
            'Business Economics': 'BE',
            'Decision Sciences': 'DS-I',
            'Financial Reporting': 'FRA',
            'Introduction to Business Analytics': 'IBA',
            'Managerial Communication': 'MC',
            'Managerial Computing': 'MComp',
            'Marketing Management': 'MM',
            'Organizational Behaviour': 'OBD',
        }
        short = next((v for k, v in short_map.items() if k.lower() in name.lower()), name[:4])

        existing = Course.query.filter_by(code=code).first()
        if not existing:
            course = Course(
                code=code,
                name=name,
                credits=credits,
                area=area,
                faculty=faculty,
                short_name=short,
                color=COURSE_COLORS[course_color_idx % len(COURSE_COLORS)]
            )
            db.session.add(course)
            db.session.flush()
            course_color_idx += 1
            courses_by_short[short] = course
        else:
            courses_by_short[existing.short_name] = existing

    db.session.commit()

    # Refresh courses_by_short from DB
    for c in Course.query.all():
        courses_by_short[c.short_name] = c

    # Column index → slot number
    col_to_slot = {2: 1, 3: 2, 5: 3, 6: 4}  # 0-indexed cols

    # Delete existing sessions before re-import
    ClassSession.query.delete()

    for row in ws.iter_rows(min_row=4, max_row=59, values_only=True):
        if row[0] is None:
            continue
        if not isinstance(row[0], datetime):
            continue

        date_val = row[0].date()
        day_name = str(row[1]).strip() if row[1] else ''

        for col_idx, slot_num in col_to_slot.items():
            cell_val = row[col_idx] if col_idx < len(row) else None
            if not cell_val:
                continue
            text = str(cell_val).strip()
            if not text:
                continue

            # Check special
            is_special = any(kw in text.upper() for kw in ['TERM', 'HOLIDAY', 'INDEPENDENCE', 'MILAD', 'BREAK', 'MID'])

            # Find matching course
            abbr = parse_subject_abbr(text)
            course_obj = courses_by_short.get(abbr) if abbr else None

            session_obj = ClassSession(
                date=date_val,
                day_name=day_name,
                slot=slot_num,
                subject_raw=text,
                course_id=course_obj.id if course_obj else None,
                is_special=is_special,
            )
            db.session.add(session_obj)

    db.session.commit()


# ─── AUTH ROUTES ───────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user, remember=True)
            if request.is_json:
                return jsonify({'success': True, 'is_admin': user.is_admin})
            return redirect(url_for('index'))
        if request.is_json:
            return jsonify({'success': False, 'error': 'Invalid credentials'}), 401
        flash('Invalid username or password', 'error')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')

        if User.query.filter_by(username=username).first():
            msg = 'Username already taken'
            return jsonify({'success': False, 'error': msg}), 400 if request.is_json else flash(msg, 'error')
        if User.query.filter_by(email=email).first():
            msg = 'Email already registered'
            return jsonify({'success': False, 'error': msg}), 400 if request.is_json else flash(msg, 'error')

        user = User(username=username, email=email)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        login_user(user, remember=True)
        if request.is_json:
            return jsonify({'success': True})
        return redirect(url_for('index'))
    return render_template('login.html', mode='register')


# ─── MAIN ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=current_user, slots=SLOTS)


@app.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        flash('Admin access required', 'error')
        return redirect(url_for('index'))
    courses = Course.query.all()
    return render_template('admin.html', user=current_user, slots=SLOTS, courses=courses)

@app.route('health')
def health():
    return jsonify({'status': 'ok', 'database': 'connected'})
# ─── API ROUTES ────────────────────────────────────────────────────────────────

@app.route('/api/today')
@login_required
def api_today():
    today = date.today()
    sessions = ClassSession.query.filter_by(date=today).order_by(ClassSession.slot).all()
    return jsonify({
        'date': today.isoformat(),
        'day': today.strftime('%A'),
        'sessions': [s.to_dict() for s in sessions],
        'slots': SLOTS
    })


@app.route('/api/sessions')
@login_required
def api_sessions():
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    try:
        start = date.fromisoformat(start_str) if start_str else date.today()
        end = date.fromisoformat(end_str) if end_str else start
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    sessions = ClassSession.query.filter(
        ClassSession.date >= start,
        ClassSession.date <= end
    ).order_by(ClassSession.date, ClassSession.slot).all()
    return jsonify([s.to_dict() for s in sessions])


@app.route('/api/sessions/week')
@login_required
def api_sessions_week():
    from datetime import timedelta
    today = date.today()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    start_str = request.args.get('start', start.isoformat())
    end_str = request.args.get('end', end.isoformat())
    try:
        start = date.fromisoformat(start_str)
        end = date.fromisoformat(end_str)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    sessions = ClassSession.query.filter(
        ClassSession.date >= start,
        ClassSession.date <= end
    ).order_by(ClassSession.date, ClassSession.slot).all()
    return jsonify({
        'start': start.isoformat(),
        'end': end.isoformat(),
        'sessions': [s.to_dict() for s in sessions]
    })


@app.route('/api/courses')
@login_required
def api_courses():
    courses = Course.query.all()
    return jsonify([c.to_dict() for c in courses])


@app.route('/api/sessions/<int:session_id>', methods=['PUT'])
@login_required
def update_session(session_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    session_obj = db.session.get(ClassSession, session_id)
    if not session_obj:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json()
    if 'subject_raw' in data:
        session_obj.subject_raw = data['subject_raw']
    if 'course_id' in data:
        session_obj.course_id = data['course_id'] or None
    if 'is_special' in data:
        session_obj.is_special = data['is_special']
    if 'notes' in data:
        session_obj.notes = data['notes']
    session_obj.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(session_obj.to_dict())


@app.route('/api/sessions', methods=['POST'])
@login_required
def create_session():
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    try:
        date_val = date.fromisoformat(data['date'])
    except (KeyError, ValueError):
        return jsonify({'error': 'Invalid date'}), 400

    # Validate slot
    slot = int(data.get('slot', 1))
    if slot not in SLOTS:
        return jsonify({'error': 'Invalid slot'}), 400

    # Check if session already exists
    existing = ClassSession.query.filter_by(date=date_val, slot=slot).first()
    if existing:
        return jsonify({'error': 'Session already exists for this date and slot'}), 409

    session_obj = ClassSession(
        date=date_val,
        day_name=date_val.strftime('%A'),
        slot=slot,
        subject_raw=data.get('subject_raw', ''),
        course_id=data.get('course_id') or None,
        is_special=data.get('is_special', False),
        notes=data.get('notes', '')
    )
    db.session.add(session_obj)
    db.session.commit()
    return jsonify(session_obj.to_dict()), 201


@app.route('/api/sessions/<int:session_id>', methods=['DELETE'])
@login_required
def delete_session(session_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    session_obj = db.session.get(ClassSession, session_id)
    if not session_obj:
        return jsonify({'error': 'Not found'}), 404
    db.session.delete(session_obj)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/courses', methods=['POST'])
@login_required
def create_course():
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    course = Course(
        code=data.get('code', ''),
        name=data.get('name', ''),
        credits=data.get('credits', 3.0),
        area=data.get('area', ''),
        faculty=data.get('faculty', ''),
        short_name=data.get('short_name', ''),
        color=data.get('color', '#6366f1'),
    )
    db.session.add(course)
    db.session.commit()
    return jsonify(course.to_dict()), 201


@app.route('/api/courses/<int:course_id>', methods=['PUT'])
@login_required
def update_course(course_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    course = db.session.get(Course, course_id)
    if not course:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json()
    for field in ['code', 'name', 'credits', 'area', 'faculty', 'short_name', 'color']:
        if field in data:
            setattr(course, field, data[field])
    db.session.commit()
    return jsonify(course.to_dict())


@app.route('/api/courses/<int:course_id>', methods=['DELETE'])
@login_required
def delete_course(course_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    course = db.session.get(Course, course_id)
    if not course:
        return jsonify({'error': 'Not found'}), 404
    db.session.delete(course)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/notifications/subscribe', methods=['POST'])
@login_required
def subscribe_notifications():
    data = request.get_json()
    notif = Notification.query.filter_by(user_id=current_user.id).first()
    if not notif:
        notif = Notification(user_id=current_user.id)
        db.session.add(notif)
    notif.push_subscription = json.dumps(data.get('subscription'))
    notif.notify_before_class = data.get('notify_before_class', True)
    notif.notify_minutes_before = int(data.get('notify_minutes_before', 15))
    notif.notify_morning = data.get('notify_morning', True)
    notif.morning_time = data.get('morning_time', '07:00')
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/notifications/settings', methods=['GET'])
@login_required
def get_notification_settings():
    notif = Notification.query.filter_by(user_id=current_user.id).first()
    if not notif:
        return jsonify({
            'notify_before_class': True,
            'notify_minutes_before': 15,
            'notify_morning': True,
            'morning_time': '07:00',
            'subscribed': False
        })
    return jsonify({
        'notify_before_class': notif.notify_before_class,
        'notify_minutes_before': notif.notify_minutes_before,
        'notify_morning': notif.notify_morning,
        'morning_time': notif.morning_time,
        'subscribed': bool(notif.push_subscription)
    })


@app.route('/api/admin/import-excel', methods=['POST'])
@login_required
def admin_import_excel():
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    excel_path = os.path.join(os.path.dirname(__file__), 'MBA BA 2026-28 Term I Schedule .xlsx')
    if not os.path.exists(excel_path):
        return jsonify({'error': 'Excel file not found'}), 404
    try:
        import_excel(excel_path)
        return jsonify({'success': True, 'message': 'Excel imported successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'is_admin': u.is_admin,
        'created_at': u.created_at.isoformat()
    } for u in users])


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Admin only'}), 403
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json()
    if 'is_admin' in data:
        user.is_admin = data['is_admin']
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/me')
@login_required
def api_me():
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'email': current_user.email,
        'is_admin': current_user.is_admin,
    })


# ─── PWA ROUTES ────────────────────────────────────────────────────────────────

@app.route('/manifest.json')
def manifest():
    from flask import send_from_directory
    return send_from_directory('static', 'manifest.json')


@app.route('/sw.js')
def service_worker():
    from flask import send_from_directory
    response = send_from_directory('static', 'sw.js')
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Service-Worker-Allowed'] = '/'
    return response


# ─── INIT DB ────────────────────────────────────────────────────────────────────

def create_default_admin():
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin', email='admin@iimsambalpur.ac.in', is_admin=True)
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print('[OK] Default admin created: admin / admin123')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'

    with app.app_context():
        db.create_all()
        create_default_admin()

        # Auto-import Excel on first run
        excel_path = os.path.join(os.path.dirname(__file__), 'MBA BA 2026-28 Term I Schedule .xlsx')
        if os.path.exists(excel_path) and ClassSession.query.count() == 0:
            print('[INFO] Importing timetable from Excel...')
            import_excel(excel_path)
            print('[OK] Timetable imported successfully!')

    app.run(debug=debug, host='0.0.0.0', port=port)
